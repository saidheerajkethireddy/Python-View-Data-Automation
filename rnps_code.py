import pandas as pd
import sys
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

CSV_PATH    = "Q1'26_RNPS_View_Data.csv"
OUTPUT_PATH = r"CNI_Desjardins_Output - Q1'26.xlsx"   # ← PASTE YOUR FULL PATH HERE

COLS = [
    "provider_name", "nps_period", "nps_nps_category_value",
    "nps_nps_weight", "nps_region", "nps_product_ownership"
]

ALL_REGIONS = [
    "Alberta", "British Columbia", "Manitoba", "New Brunswick",
    "Newfoundland and Labrador", "Nova Scotia",
    "Nunavut / Northwest Territories", "Ontario",
    "Prince Edward Island", "Quebec", "Saskatchewan", "Yukon"
]

REGION_OPTIONS = {
    1: ("All of Canada",  ALL_REGIONS),
    2: ("Quebec",         ["Quebec"]),
    3: ("Rest of Canada", [r for r in ALL_REGIONS if r != "Quebec"]),
}

PROVIDER_DISPLAY_NAMES = {
    "CAA Insurance Company":          "CAA",
    "The Personal":                   "The Personal",
    "Desjardins Agent Network":       "DAN",
    "Co-operators":                   "Co-operators",
    "Allstate":                       "Allstate",
    "DI Direct":                      "DI Direct",
    "RBC Insurance":                  "RBC",
    "Intact Insurance":               "Intact",
    "BelairDirect":                   "BelairDirect",
    "Aviva":                          "Aviva",
    "TD Insurance":                   "TD Insurance",
    "Wawanesa Insurance":             "Wawanesa",
    "Promutuel Assurance":            "Promutuel",
    "Beneva":                         "Beneva",
    "Industrial Alliance (iA)":       "iA",
    "Definity":                       "Definity",
    "Economical (becoming Definity)": "Economical (becoming Definity)",
}

REGION_CONFIG = {
    1: {  # All of Canada
        "providers": [
            "CAA Insurance Company", "The Personal", "Desjardins Agent Network",
            "Co-operators", "Allstate", "DI Direct",
            "RBC Insurance", "Intact Insurance", "BelairDirect",
            "Aviva", "TD Insurance", "Wawanesa Insurance",
            "Promutuel Assurance", "Beneva", "Industrial Alliance (iA)",
            "Definity", "Economical (becoming Definity)",
        ],
    },
    2: {  # Quebec
        "providers": [
            "The Personal", "Promutuel Assurance", "DI Direct",
            "Beneva", "Industrial Alliance (iA)", "Intact Insurance",
            "BelairDirect", "CAA Insurance Company",
            "Definity", "Economical (becoming Definity)",
        ],
    },
    3: {  # Rest of Canada
        "providers": [
            "CAA Insurance Company", "The Personal", "Desjardins Agent Network",
            "Co-operators", "Allstate", "DI Direct",
            "RBC Insurance", "Intact Insurance", "BelairDirect",
            "Aviva", "TD Insurance", "Wawanesa Insurance",
            "Definity", "Economical (becoming Definity)",
        ],
    },
}

PERIOD_ORDER = [
    "2020", "Q2 2021", "Q3+Q4 2021",
    "Q1 2022", "Q2 2022", "Q3 2022", "Q4 2022",
    "Q1 2023", "Q2 2023", "Q3 2023", "Q4 2023",
    "Q1 2024", "Q2 2024", "Q3 2024", "Q4 2024",
    "Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025"
]

ALL_PRODUCTS = ["Auto", "Home", "Auto + Home"]
N_THRESHOLD  = 50

# ── Utilities ──────────────────────────────────────────────────────────────────

def sep(char="─", w=72): print(char * w)
def sep2(w=72):          print("═" * w)

def header(title):
    sep2()
    print(f"  {title}")
    sep2()

def print_numbered(items, cols=2):
    if not items: return
    rows = -(-len(items) // cols)
    col_w = max(len(f"  {i+1:2}. {item}") for i, item in enumerate(items)) + 4
    for r in range(rows):
        line = ""
        for c in range(cols):
            idx = r + c * rows
            if idx < len(items):
                entry = f"  {idx+1:2}. {items[idx]}"
                line += entry.ljust(col_w)
        print(line)

def parse_selection(raw, max_n, allow_all=True, exactly_one=False):
    raw = raw.strip()
    if allow_all and raw.lower() == "all":
        return list(range(max_n))
    tokens = raw.replace(",", " ").split()
    indices, seen = [], set()
    for t in tokens:
        try:
            n = int(t)
            if n < 1 or n > max_n:
                print(f"  ⚠️  '{n}' out of range (1–{max_n}). Try again.")
                return None
            if n - 1 not in seen:
                indices.append(n - 1)
                seen.add(n - 1)
        except ValueError:
            print(f"  ⚠️  '{t}' is not a number. Try again.")
            return None
    if not indices:
        print("  ⚠️  No selection made. Try again.")
        return None
    if exactly_one and len(indices) != 1:
        print("  ⚠️  Please select exactly ONE. Try again.")
        return None
    return indices

def ask(question):
    print(f"\n{question}")
    return input("  >>> ").strip()

# ── Rolling period logic ───────────────────────────────────────────────────────

def get_rolling_periods(anchor, n_quarters, avail_periods):
    ordered = [p for p in PERIOD_ORDER if p in avail_periods]
    if anchor not in ordered:
        return [anchor]
    idx   = ordered.index(anchor)
    start = max(0, idx - n_quarters + 1)
    return ordered[start: idx + 1]

# ── rNPS formula ───────────────────────────────────────────────────────────────

def compute_rnps(data):
    rows = []
    for provider, grp in data.groupby("provider_name"):
        num   = (grp["nps_nps_weight"] * grp["nps_nps_category_value"]).sum()
        denom = grp["nps_nps_weight"].sum()
        rnps  = (num / denom) * 100 if denom != 0 else None
        rows.append({"Provider": provider, "rNPS": rnps, "N": len(grp)})
    return pd.DataFrame(rows).set_index("Provider")

def compute_rolling_rnps(df2, periods):
    return compute_rnps(df2[df2["nps_period"].isin(periods)])

def weighted_avg(raw_df):
    """Single weighted rNPS across all respondents in raw_df."""
    if raw_df is None or len(raw_df) == 0:
        return None
    num   = (raw_df["nps_nps_weight"] * raw_df["nps_nps_category_value"]).sum()
    denom = raw_df["nps_nps_weight"].sum()
    return (num / denom) * 100 if denom != 0 else None

# ── Main output: summary + table, then save ───────────────────────────────────

def print_and_save(results, ctx):
    region_key = ctx["region_key"]
    display    = REGION_CONFIG[region_key]["providers"]
    def dname(p): return PROVIDER_DISPLAY_NAMES.get(p, p)

    # ── Get start cell from ctx ───────────────────────────────────────────────
    cell_input = ctx["cell_input"]
    m = re.match(r"^([A-Z]+)([0-9]+)$", cell_input)
    start_col  = column_index_from_string(m.group(1))
    start_row  = int(m.group(2))

    # ── Build values ───────────────────────────────────────────────────────────
    def fval(df, p, metric):
        if p in df.index and df.loc[p, metric] is not None:
            v = df.loc[p, metric]
            return str(int(v)) if metric == "N" else f"{v:.4f}"
        return "—"

    raw_dfs = ctx.get("raw_dfs", {})
    ind_avgs = {label: weighted_avg(raw_dfs.get(label)) for label, _ in results}

    # ── Print summary header ───────────────────────────────────────────────────
    comp_type = ctx.get("comp_type", "")
    cur  = ctx['cur_label']
    prev = ctx['comp_label'] if ctx.get('comp_label') else "—"
    reg  = ctx['region_label']
    print(f"\n  rNPS | Cur: {cur} | Prev({comp_type}): {prev} | {reg}")

    # ── Terminal table: periods as rows, providers as columns ────────────────
    # Columns: providers + Industry Avg
    col_headers = [dname(p) for p in display] + ["Industry Avg"]
    LABEL_W = 26
    COL_W   = max(max(len(h) for h in col_headers), 10) + 2
    total_w = LABEL_W + COL_W * len(col_headers)

    print()
    # Provider header row
    print(" " * LABEL_W, end="")
    for h in col_headers:
        print(f"{h:>{COL_W}}", end="")
    print()
    sep(w=total_w)

    # rNPS rows (one per result)
    for label, df in results:
        print(f"  {label + ' rNPS':<{LABEL_W - 2}}", end="")
        for p in display:
            v = df.loc[p, "rNPS"] if p in df.index and df.loc[p, "rNPS"] is not None else None
            print(f"{f'{v:.4f}' if v is not None else '—':>{COL_W}}", end="")
        a = ind_avgs.get(label)
        print(f"{f'{a:.4f}' if a is not None else '—':>{COL_W}}")

    sep(w=total_w)

    # N row — current period only
    label0, df0 = results[0]
    print(f"  {label0 + ' N':<{LABEL_W - 2}}", end="")
    for p in display:
        v = df0.loc[p, "N"] if p in df0.index else None
        print(f"{str(int(v)) if v is not None else '—':>{COL_W}}", end="")
    print(f"{'—':>{COL_W}}")
    sep2(w=total_w)

    # ── Excel ──────────────────────────────────────────────────────────────────
    try:
        wb = load_workbook(OUTPUT_PATH)
    except FileNotFoundError:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    DATA_SHEET    = "Data"
    SUMMARY_SHEET = "Summary"
    wd     = wb[DATA_SHEET]     if DATA_SHEET    in wb.sheetnames else wb.create_sheet(title=DATA_SHEET)
    ws_sum = wb[SUMMARY_SHEET]  if SUMMARY_SHEET in wb.sheetnames else None

    if ws_sum is None:
        ws_sum = wb.create_sheet(title=SUMMARY_SHEET)
        sum_headers = ["Timestamp", "Rolling Window", "Region", "Product(s)",
                       "Current Period", "Periods Included",
                       "Comparison Period", "Comp. Periods Incl.",
                       "N Threshold", "Start Cell", "Providers"]
        for ci, h in enumerate(sum_headers, start=1):
            hc = ws_sum.cell(row=1, column=ci, value=h)
            hc.font      = Font(name="Arial", bold=True, size=10)
            hc.fill      = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")
            hc.alignment = Alignment(horizontal="center", vertical="center")
        ws_sum.row_dimensions[1].height = 18

    GREY_FILL  = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")
    RED_FILL   = PatternFill("solid", start_color="FFCCCC", end_color="FFCCCC")
    RED_HDR    = PatternFill("solid", start_color="FF4444", end_color="FF4444")
    HDR_FONT   = Font(name="Arial", bold=True, size=10)
    DATA_FONT  = Font(name="Arial", size=10)
    WHITE_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    CENTER     = Alignment(horizontal="center", vertical="center")
    LEFT       = Alignment(horizontal="left",   vertical="center")
    thin       = Side(style="thin", color="AAAAAA")
    BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style(cell, font=None, fill=None, align=None, num_fmt=None):
        if font:    cell.font          = font
        if fill:    cell.fill          = fill
        if align:   cell.alignment     = align
        if num_fmt: cell.number_format = num_fmt
        cell.border = BORDER

    # Build data rows
    data_rows = []
    for label, df in results:
        row = [f"{label} rNPS"]
        for p in display:
            v = df.loc[p, "rNPS"] if p in df.index and df.loc[p, "rNPS"] is not None else None
            row.append(round(float(v), 4) if v is not None else "")
        a = ind_avgs.get(label)
        row.append(round(a, 4) if a is not None else "")
        data_rows.append(row)

    label0, df0 = results[0]
    nrow = [f"{label0} N"]
    for p in display:
        v = df0.loc[p, "N"] if p in df0.index else None
        nrow.append(int(v) if v is not None else "")
    nrow.append("")
    data_rows.append(nrow)

    avg_col = start_col + len(display) + 1

    # Summary label — one row above the table
    comp_type = ctx.get("comp_type", "")
    cur_lbl   = ctx['cur_label']
    prev_lbl  = ctx['comp_label'] if ctx.get('comp_label') else "—"
    reg_lbl   = ctx['region_label']
    summary   = f"rNPS | Cur: {cur_lbl} | Prev({comp_type}): {prev_lbl} | {reg_lbl}"
    sum_cell  = wd.cell(row=start_row - 1, column=start_col, value=summary)
    sum_cell.font      = Font(name="Arial", bold=True, size=10)
    sum_cell.alignment = Alignment(horizontal="left", vertical="center")
    wd.row_dimensions[start_row - 1].height = 16

    # Header row
    style(wd.cell(row=start_row, column=start_col, value=""), font=HDR_FONT, fill=GREY_FILL, align=LEFT)
    for ci, provider in enumerate(display, start=start_col + 1):
        style(wd.cell(row=start_row, column=ci, value=dname(provider)), font=HDR_FONT, fill=GREY_FILL, align=CENTER)
    style(wd.cell(row=start_row, column=avg_col, value="Industry Avg"), font=WHITE_FONT, fill=RED_HDR, align=CENTER)
    wd.row_dimensions[start_row].height = 18

    # Data rows
    for ri, row in enumerate(data_rows, start=start_row + 1):
        is_n = row[0].endswith(" N")
        style(wd.cell(row=ri, column=start_col, value=row[0]), font=HDR_FONT, fill=GREY_FILL, align=LEFT)
        for ci, val in enumerate(row[1:-1], start=start_col + 1):
            style(wd.cell(row=ri, column=ci, value=val), font=DATA_FONT, align=CENTER,
                  num_fmt="0" if is_n else "0.0000")
        style(wd.cell(row=ri, column=avg_col, value=row[-1]),
              font=Font(name="Arial", bold=True, size=10),
              fill=RED_FILL, align=CENTER, num_fmt="" if is_n else "0.0000")
        wd.row_dimensions[ri].height = 18

    wd.column_dimensions[get_column_letter(start_col)].width = 32
    for ci in range(start_col + 1, avg_col):
        wd.column_dimensions[get_column_letter(ci)].width = 14
    wd.column_dimensions[get_column_letter(avg_col)].width = 16

    # Summary row
    next_row = ws_sum.max_row + 1
    summary_vals = [
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        ctx.get("r_label", ""), ctx.get("region_label", ""), ctx.get("products", ""),
        ctx.get("cur_label", ""), ctx.get("cur_periods_str", ""),
        ctx.get("comp_label") or "", ctx.get("comp_periods_str") or "",
        f"N > {N_THRESHOLD}", cell_input, ", ".join(dname(p) for p in display),
    ]
    for ci, (val, w) in enumerate(zip(summary_vals, [18,14,16,22,16,40,16,40,12,10,80]), start=1):
        sc = ws_sum.cell(row=next_row, column=ci, value=val)
        sc.font = DATA_FONT; sc.alignment = LEFT
        ws_sum.column_dimensions[get_column_letter(ci)].width = w
    ws_sum.row_dimensions[next_row].height = 16

    wb.save(OUTPUT_PATH)
    print(f"\n  ✅  Saved → {OUTPUT_PATH}  |  Sheet '{DATA_SHEET}' at {cell_input}  |  Summary row {next_row}\n")

# ── Wizard ─────────────────────────────────────────────────────────────────────

# ── 6 blocks config: (region_key, comp_type, cell)
# comp_type: "yoy" = 4 quarters back, "qoq" = 1 quarter back
BLOCKS = [
    (1, "yoy", "B108"),   # All of Canada   — YoY
    (2, "yoy", "B118"),   # Quebec          — YoY
    (3, "yoy", "O118"),   # Rest of Canada  — YoY
    (1, "qoq", "B128"),   # All of Canada   — QoQ
    (2, "qoq", "B138"),   # Quebec          — QoQ
    (3, "qoq", "O138"),   # Rest of Canada  — QoQ
]

R_SIZE  = 4               # always Rolling 4
PRODUCTS = ALL_PRODUCTS   # always all products


def run_block(df, region_key, comp_type, cell_input):
    region_label, selected_regions = REGION_OPTIONS[region_key]
    r_label = f"R{R_SIZE}"

    mask  = df["nps_region"].isin(selected_regions) & df["nps_product_ownership"].isin(PRODUCTS)
    df2   = df[mask].copy()
    avail = [p for p in PERIOD_ORDER if p in df2["nps_period"].unique()]

    cur_anchor = avail[-1]
    back       = 4 if comp_type == "yoy" else 1
    comp_idx   = avail.index(cur_anchor) - back
    comp_anchor = avail[comp_idx] if comp_idx >= 0 else None

    cur_periods  = get_rolling_periods(cur_anchor,  R_SIZE, avail)
    comp_periods = get_rolling_periods(comp_anchor, R_SIZE, avail) if comp_anchor else []

    cur_label  = f"{r_label} {cur_anchor}"
    comp_label = f"{r_label} {comp_anchor}" if comp_anchor else None

    results = [(cur_label, compute_rolling_rnps(df2, cur_periods))]
    if comp_anchor:
        results.append((comp_label, compute_rolling_rnps(df2, comp_periods)))

    df_all = df[df["nps_region"].isin(selected_regions)].copy()
    df_all["nps_nps_category_value"] = pd.to_numeric(df_all["nps_nps_category_value"], errors="coerce")
    df_all["nps_nps_weight"]         = pd.to_numeric(df_all["nps_nps_weight"],         errors="coerce")
    df_all = df_all.dropna(subset=["nps_nps_category_value","nps_nps_weight"])
    raw_dfs = {cur_label: df_all[df_all["nps_period"].isin(cur_periods)].copy()}
    if comp_anchor:
        raw_dfs[comp_label] = df_all[df_all["nps_period"].isin(comp_periods)].copy()

    ctx = {
        "r_label":          r_label,
        "region_key":       region_key,
        "region_label":     region_label,
        "comp_type":        comp_type.upper(),
        "products":         ", ".join(PRODUCTS),
        "cur_label":        cur_label,
        "cur_periods_str":  ", ".join(cur_periods),
        "comp_label":       comp_label,
        "comp_periods_str": ", ".join(comp_periods) if comp_periods else None,
        "raw_dfs":          raw_dfs,
        "cell_input":       cell_input,
    }
    print_and_save(results, ctx)


def main():
    header("rNPS CALCULATOR  —  Canada Insurance")
    print("\n📂  Loading data...")
    df = pd.read_csv(CSV_PATH, usecols=COLS, low_memory=False)
    df["nps_nps_category_value"] = pd.to_numeric(df["nps_nps_category_value"], errors="coerce")
    df["nps_nps_weight"]         = pd.to_numeric(df["nps_nps_weight"],         errors="coerce")
    df = df.dropna(subset=["nps_nps_category_value", "nps_nps_weight"])
    print("✅  Ready.\n")

    try:
        for region_key, comp_type, cell in BLOCKS:
            run_block(df, region_key, comp_type, cell)
    except (KeyboardInterrupt, EOFError):
        print("\n\nInterrupted.")
        sys.exit()

    print("\n✅  All 6 blocks done.")

if __name__ == "__main__":
    main()